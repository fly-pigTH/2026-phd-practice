#!/usr/bin/env python3
import csv
import re
import unicodedata
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
PDF_PATH = ROOT / "data" / "raw" / "项目申请 -清华大学研究生社会实践系统.pdf"
OUT_DIR = ROOT / "data" / "processed"
XLSX_PATH = OUT_DIR / "项目申请信息表.xlsx"
CSV_PATH = OUT_DIR / "项目申请信息表.csv"


EXTRA_CHAR_MAP = str.maketrans(
    {
        "⻘": "青",
        "⻓": "长",
        "⻮": "齿",
        "⻩": "黄",
        "⻢": "马",
        "⻛": "风",
        "⻋": "车",
        "⻚": "页",
        "⻆": "角",
        "⻜": "飞",
        "⻰": "龙",
        "⻣": "骨",
        "⻧": "卤",
        "⻬": "齐",
        "⻥": "鱼",
        "⻄": "西",
        "⻅": "见",
        "⻔": "门",
        "⻝": "食",
        "⻉": "贝",
        "⻉": "贝",
        "⼀": "一",
        "⼆": "二",
        "⼈": "人",
        "⼤": "大",
        "⼯": "工",
        "⼝": "口",
        "⼭": "山",
        "⼴": "广",
        "⼼": "心",
        "⽅": "方",
        "⽆": "无",
        "⽣": "生",
        "⽤": "用",
        "⽬": "目",
        "⽴": "立",
        "⽹": "网",
        "⾏": "行",
        "⾦": "金",
        "⾯": "面",
        "⾸": "首",
        "⺠": "民",
    }
)


def normalize_text(value: str) -> str:
    value = unicodedata.normalize("NFKC", value).translate(EXTRA_CHAR_MAP)
    value = value.replace("\u200b", "").replace("\ufeff", "")
    value = re.sub(r"[ \t]+", " ", value)
    return value.strip()


def useful_line(line: str) -> bool:
    if not line:
        return False
    skip_prefixes = (
        "2026/4/30",
        "首页",
        "已提交申请",
        "项目查询",
        "项目名称11",
        "需求院系",
        "标签:",
        "查询 重置",
        "满足条件记录合计",
        "https://",
    )
    return (
        not line.startswith(skip_prefixes)
        and "第一阶段选课" not in line
        and "朱颖雷" not in line
    )


def extract_lines() -> list[str]:
    lines: list[str] = []
    with pdfplumber.open(PDF_PATH) as pdf:
        for page_no, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
            for raw_line in text.splitlines():
                line = normalize_text(raw_line)
                if useful_line(line):
                    lines.append(line)
    return lines


BASE_RE = re.compile(
    r"基地信息:(?P<base>.*?)\s+(?P<level>[校系]级基地)\s+.*?"
    r"(?P<province>[\u4e00-\u9fff]+省|[\u4e00-\u9fff]+自治区|北京市|上海市|天津市|重庆市|香港|澳门|台湾|无|null)"
)


def parse_projects(lines: list[str]) -> list[dict[str, object]]:
    projects: list[dict[str, object]] = []
    pending_title_parts: list[str] = []
    pending: dict[str, object] | None = None

    for line in lines:
        if "需求人数:" in line:
            title_fragment, rest = line.split("需求人数:", 1)
            title = normalize_text(" ".join(pending_title_parts + [title_fragment]))
            pending_title_parts = []
            is_key = title.startswith("重点 ")
            if is_key:
                title = title[3:].strip()
            need_match = re.search(r"^(\d+)", rest)
            preferences = {name: "" for name in ("第一", "第二", "第三", "第四")}
            for name, count in re.findall(r"(第一|第二|第三|第四)志愿[:：](\d+)", rest):
                preferences[name] = int(count)
            note = ""
            if "已申请人数" not in line:
                note = "原PDF未显示志愿申请数"
            elif preferences["第四"] == "":
                note = "PDF右侧截断，第四志愿数未显示"
            pending = {
                "序号": len(projects) + 1,
                "是否重点": "是" if is_key else "否",
                "项目名称": title,
                "需求人数": int(need_match.group(1)) if need_match else "",
                "第一志愿申请数": preferences["第一"],
                "第二志愿申请数": preferences["第二"],
                "第三志愿申请数": preferences["第三"],
                "第四志愿申请数": preferences["第四"],
                "基地名称": "",
                "基地级别": "",
                "基地省市": "",
                "备注": note,
            }
            continue

        if pending and line.startswith("基地信息:"):
            base = BASE_RE.search(line)
            if base:
                pending["基地名称"] = normalize_text(base.group("base"))
                pending["基地级别"] = base.group("level")
                pending["基地省市"] = "" if base.group("province") == "null" else base.group("province")
            else:
                pending["基地名称"] = normalize_text(line.removeprefix("基地信息:"))
            projects.append(pending)
            pending = None
            continue

        if pending is None and not line.startswith("基地信息:"):
            pending_title_parts.append(line)

    return projects


def write_csv(projects: list[dict[str, object]], headers: list[str]) -> None:
    OUT_DIR.mkdir(exist_ok=True)
    with CSV_PATH.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(projects)


def write_xlsx(projects: list[dict[str, object]], headers: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "项目申请信息"
    ws.append(headers)
    for project in projects:
        ws.append([project.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "A": 8,
        "B": 10,
        "C": 70,
        "D": 10,
        "E": 14,
        "F": 14,
        "G": 14,
        "H": 14,
        "I": 46,
        "J": 12,
        "K": 12,
        "L": 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    table = Table(displayName="ProjectApplications", ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    wb.save(XLSX_PATH)


def main() -> None:
    headers = [
        "序号",
        "是否重点",
        "项目名称",
        "需求人数",
        "第一志愿申请数",
        "第二志愿申请数",
        "第三志愿申请数",
        "第四志愿申请数",
        "基地名称",
        "基地级别",
        "基地省市",
        "备注",
    ]
    lines = extract_lines()
    projects = parse_projects(lines)
    write_csv(projects, headers)
    write_xlsx(projects, headers)
    print(f"lines={len(lines)}")
    print(f"projects={len(projects)}")
    print(f"xlsx={XLSX_PATH}")
    print(f"csv={CSV_PATH}")


if __name__ == "__main__":
    main()
