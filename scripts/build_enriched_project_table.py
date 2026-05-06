#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from build_dashboard import REGION_MAP, infer_province, to_int


ROOT = Path(__file__).resolve().parents[1]
MAIN_CSV = ROOT / "data" / "processed" / "项目申请信息表.csv"
DETAIL_JSON = ROOT / "data" / "processed" / "项目详情信息表.json"
OUT_CSV = ROOT / "data" / "processed" / "项目完整信息表.csv"
OUT_XLSX = ROOT / "data" / "processed" / "项目完整信息表.xlsx"

DETAIL_FIELDS = [
    "项目ID",
    "详情链接",
    "项目编号",
    "所属基地编号",
    "研究方向",
    "预计工作量(天)",
    "所需准备工作",
    "单位名称",
    "单位省市",
    "单位区县",
    "单位地址",
    "项目背景",
    "项目目标",
    "需要解决的关键技术问题",
    "现有条件",
    "时间安排",
    "详情介绍",
    "HTML文件",
]

DERIVED_FIELDS = [
    "项目省市(用于地图)",
    "地域分区",
    "省市来源",
    "申请数据状态",
    "申请数合计(已知)",
    "竞争度(已知/下限)",
]


def load_details() -> dict[str, dict[str, str]]:
    rows = json.loads(DETAIL_JSON.read_text(encoding="utf-8"))
    return {str(row["序号"]): row for row in rows}


def build_rows() -> tuple[list[str], list[dict[str, str]]]:
    details = load_details()
    with MAIN_CSV.open(encoding="utf-8-sig") as f:
        main_rows = list(csv.DictReader(f))

    headers = list(main_rows[0].keys()) + DERIVED_FIELDS + DETAIL_FIELDS
    rows: list[dict[str, str]] = []
    for main in main_rows:
        detail = details.get(main["序号"], {})
        merged = dict(main)
        # Detail pages have the untruncated canonical project title.
        if detail.get("项目名称"):
            merged["项目名称"] = detail["项目名称"]
        province = infer_province(main["基地省市"], main["基地名称"])
        source = "原始字段" if main["基地省市"].strip() else ("规则推断" if province != "未注明" else "未注明")
        if province == "未注明" and detail.get("单位省市"):
            province = detail["单位省市"].strip()
            source = "单位省市推断"
        known_total = sum(to_int(main.get(field, "")) for field in ["第一志愿申请数", "第二志愿申请数", "第三志愿申请数", "第四志愿申请数"])
        note = main["备注"].strip()
        if note == "原PDF未显示志愿申请数":
            apply_status = "申请数缺失"
            competition = ""
        elif note == "PDF右侧截断，第四志愿数未显示":
            apply_status = "第四志愿缺失"
            competition = f">={known_total / max(to_int(main['需求人数']), 1):.2f}"
        else:
            apply_status = "完整"
            competition = f"{known_total / max(to_int(main['需求人数']), 1):.2f}"
        merged.update(
            {
                "项目省市(用于地图)": province,
                "地域分区": REGION_MAP.get(province, "未分区"),
                "省市来源": source,
                "申请数据状态": apply_status,
                "申请数合计(已知)": str(known_total) if apply_status != "申请数缺失" else "",
                "竞争度(已知/下限)": competition,
            }
        )
        for field in DETAIL_FIELDS:
            merged[field] = detail.get(field, "")
        rows.append(merged)
    return headers, rows


def write_csv(headers: list[str], rows: list[dict[str, str]]) -> None:
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(headers: list[str], rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "项目完整信息"
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    width_by_header = {
        "序号": 8,
        "是否重点": 10,
        "项目名称": 72,
        "需求人数": 10,
        "基地名称": 34,
        "基地级别": 12,
        "基地省市": 16,
        "备注": 28,
        "项目省市(用于地图)": 18,
        "地域分区": 12,
        "省市来源": 12,
        "申请数据状态": 14,
        "申请数合计(已知)": 16,
        "竞争度(已知/下限)": 18,
        "详情链接": 48,
        "所需准备工作": 54,
        "单位地址": 44,
        "项目背景": 54,
        "项目目标": 64,
        "需要解决的关键技术问题": 54,
        "现有条件": 42,
        "时间安排": 42,
        "详情介绍": 62,
        "HTML文件": 44,
    }
    for index, header in enumerate(headers, start=1):
        col_letter = ws.cell(row=1, column=index).column_letter
        ws.column_dimensions[col_letter].width = width_by_header.get(header, 16)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    table = Table(displayName="EnrichedProjectInfo", ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    wb.save(OUT_XLSX)


def main() -> None:
    headers, rows = build_rows()
    write_csv(headers, rows)
    write_xlsx(headers, rows)
    print(f"rows={len(rows)}")
    print(f"csv={OUT_CSV}")
    print(f"xlsx={OUT_XLSX}")


if __name__ == "__main__":
    main()
